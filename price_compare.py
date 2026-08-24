import os
import html
import re
import io
import csv
import zipfile
import xlrd
import requests
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BTM_FILE = "BTM_export.xls"
OUTPUT_FILE = "BTM_TTE_poredjenje.xlsx"

ETAIL_KEY = os.environ.get("TTE_ETAIL_API_KEY")
HAVIT_KEY = os.environ.get("TTE_HAVIT_API_KEY")
ETAIL_URL = f"https://tte.rs/api/sr/products?output_type=xml&api_key={ETAIL_KEY}"
HAVIT_URL = f"https://tte.rs/api/sr/products?output_type=xml&api_key={HAVIT_KEY}"

EXCLUDED_BRANDS = {"bavin", "bavitel", "havit", "powerology", "green lion", "porodo"}


def clean(v):
    if v is None:
        return ""
    v = html.unescape(str(v))
    v = re.sub(r"<[^>]+>", " ", v)
    return re.sub(r"\s+", " ", v.replace("\xa0", " ").replace("\n", " ").replace("\r", " ").replace("\t", " ")).strip()


def norm_text(v):
    v = clean(v).lower()
    for a, b in {"č":"c", "ć":"c", "š":"s", "ž":"z", "đ":"d"}.items():
        v = v.replace(a, b)
    return " ".join(re.sub(r"[^a-z0-9]+", " ", v).split())


def norm_ean(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\D", "", s)


def number(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean(v).replace("RSD", "").replace("rsd", "").replace("DIN", "").replace("din", "").replace(" ", "")
    if re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
        s = s.replace(".", "")
    elif re.fullmatch(r"\d{1,3}(,\d{3})+", s):
        s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def fmt(v):
    if v is None:
        return ""
    v = round(float(v), 2)
    return int(v) if v.is_integer() else v


def find_col(headers, names):
    h = [norm_text(x) for x in headers]
    n = [norm_text(x) for x in names]
    for i, x in enumerate(h):
        if x in n:
            return i
    for i, x in enumerate(h):
        if any(y in x for y in n):
            return i
    return None


def read_export(data):
    # BTM export je UTF-16 TSV; newline="" sprečava csv grešku
    if data.startswith((b"\xff\xfe", b"\xfe\xff")):
        text = data.decode("utf-16")
    else:
        try:
            text = data.decode("utf-16")
        except UnicodeDecodeError:
            text = data.decode("utf-8-sig", errors="replace")

    first = text.splitlines()[0] if text.splitlines() else ""
    if "\t" in first:
        return list(csv.reader(io.StringIO(text, newline=""), delimiter="\t"))

    if zipfile.is_zipfile(io.BytesIO(data)):
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
        wb.close()
        return rows

    wb = xlrd.open_workbook(file_contents=data, formatting_info=False)
    ws = wb.sheet_by_index(0)
    return [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(ws.nrows)]


def header_row(rows):
    best = None
    score_best = -1
    for i in range(min(50, len(rows))):
        s = " | ".join(clean(x).lower() for x in rows[i])
        score = sum(["ean" in s or "barkod" in s, "naziv" in s, "cena" in s, "sifra" in s or "šifra" in s])
        if score > score_best:
            best, score_best = i, score
    if best is None:
        raise Exception("Ne mogu da pronađem zaglavlje BTM tabele.")
    return best


def read_btm():
    if not os.path.exists(BTM_FILE):
        raise Exception(f"BTM fajl ne postoji: {BTM_FILE}")
    with open(BTM_FILE, "rb") as f:
        data = f.read()
    rows = read_export(data)
    hr = header_row(rows)
    headers = [clean(x) for x in rows[hr]]

    ean_c = find_col(headers, ["EAN", "Barkod", "Barcode", "EAN kod"])
    name_c = find_col(headers, ["Naziv", "Naziv artikla", "Product name", "Name"])
    price_c = find_col(headers, ["Cena", "Cena RSD", "Cena bez PDV", "Cena NETO", "Neto cena", "Price"])
    brand_c = find_col(headers, ["Brend", "Brand", "Proizvođač", "Proizvodjac"])
    code_c = find_col(headers, ["TTE šifra", "TTE sifra", "Šifra", "Sifra", "Product code", "Code"])

    if ean_c is None or name_c is None or price_c is None:
        raise Exception("Ne mogu da pronađem EAN, naziv ili cenu u BTM exportu.")

    out = []
    for raw in rows[hr + 1:]:
        row = list(raw) + [""] * max(0, len(headers) - len(raw))
        ean = norm_ean(row[ean_c])
        name = clean(row[name_c])
        brand = clean(row[brand_c]) if brand_c is not None else ""
        p = number(row[price_c])
        code = clean(row[code_c]) if code_c is not None else ""
        if not ean or not name or p is None:
            continue
        bn = norm_text(brand)
        nn = norm_text(name)
        if bn in EXCLUDED_BRANDS:
            continue
        if not bn and any(re.search(r"\b" + re.escape(b) + r"\b", nn) for b in EXCLUDED_BRANDS):
            continue
        out.append({"ean": ean, "name": name, "net": p, "tte": code})
    print("BTM artikala posle filtera:", len(out))
    return out


def xml_download(url, label):
    r = requests.get(url, timeout=120, headers={"User-Agent": "Mozilla/5.0"})
    r.raise_for_status()
    if not r.content:
        raise Exception(f"{label}: XML je prazan.")
    ET.fromstring(r.content)
    return r.content


def parse_xml(data, label):
    root = ET.fromstring(data)
    out = {}
    for product in root.findall(".//product"):
        ean = norm_ean(product.findtext("ean"))
        if not ean:
            continue
        # Isključivo netto_price, bez fallback-a na price.
        p = number(product.findtext("netto_price"))
        if p is None:
            continue
        if ean not in out:
            out[ean] = {"net": p, "tte": clean(product.findtext("article_number"))}
    print(f"{label}: EAN + netto_price = {len(out)}")
    return out


def main():
    if not ETAIL_KEY:
        raise Exception("Nedostaje GitHub Secret: TTE_ETAIL_API_KEY")
    if not HAVIT_KEY:
        raise Exception("Nedostaje GitHub Secret: TTE_HAVIT_API_KEY")

    btm = read_btm()
    etail = parse_xml(xml_download(ETAIL_URL, "ETAIL SPEC"), "ETAIL SPEC")
    havit = parse_xml(xml_download(HAVIT_URL, "HAVIT"), "HAVIT")

    results = []
    for b in btm:
        h = havit.get(b["ean"])
        e = etail.get(b["ean"])
        if h is None and e is None:
            continue

        btm_p = b["net"]
        havit_p = h["net"] if h else None
        etail_p = e["net"] if e else None

        # KONAČNA LOGIKA: BTM - TTE
        # Minus = BTM je povoljniji.
        dh = round(btm_p - havit_p, 2) if havit_p is not None else None
        de = round(btm_p - etail_p, 2) if etail_p is not None else None

        if not ((dh is not None and abs(dh) >= 0.01) or (de is not None and abs(de) >= 0.01)):
            continue

        tte = (h.get("tte") if h else "") or (e.get("tte") if e else "") or b["tte"]
        results.append([b["ean"], tte, b["name"], fmt(btm_p), fmt(havit_p), fmt(dh), fmt(etail_p), fmt(de)])

    results.sort(key=lambda x: x[0])

    wb = Workbook()
    ws = wb.active
    ws.title = "BTM vs TTE"
    headers = ["EAN", "TTE šifra", "Artikal", "BTM cena NETO", "HAVIT cena NETO", "Razlika HAVIT", "ETAIL SPEC cena NETO", "Razlika ETAIL SPEC"]
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    for row in results:
        ws.append(row)
    for row in ws.iter_rows(min_row=2):
        for c in (4, 5, 6, 7, 8):
            if isinstance(row[c-1].value, (int, float)):
                row[c-1].number_format = '#,##0.00'
    for i, width in enumerate([16, 18, 60, 18, 18, 16, 21, 20], 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    wb.save(OUTPUT_FILE)

    print("ARTIKALA SA RAZLIKOM:", len(results))
    print("IZLAZ:", OUTPUT_FILE)
    print("RAZLIKA = BTM - TTE")


if __name__ == "__main__":
    main()
