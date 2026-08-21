import os
import re
import html
import xlrd
import requests
import xml.etree.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BTM_FILE = "BTM_export.xls"
OUTPUT_FILE = "BTM_TTE_poredjenje.xlsx"
PDV = 1.20

# Brendovi koji se NE porede.
EXCLUDED_BRANDS = {
    "havit",
    "jbl",
    "samsung",
    "apple",
    "powerology",
    "green lion",
    "porodo",
}

TTE_ETAIL_API_KEY = os.environ.get("TTE_ETAIL_API_KEY")

ETAIL_URL = (
    "https://tte.rs/api/sr/products"
    "?output_type=xml&api_key="
    + str(TTE_ETAIL_API_KEY)
)


def clean_text(value):
    if value is None:
        return ""
    value = str(value)
    value = html.unescape(value)
    value = re.sub(r"<[^>]+>", " ", value)
    value = value.replace("\xa0", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def normalize_text(value):
    value = clean_text(value).lower()
    replacements = {"č": "c", "ć": "c", "š": "s", "ž": "z", "đ": "d"}
    for old, new in replacements.items():
        value = value.replace(old, new)
    return " ".join(re.sub(r"[^a-z0-9]+", " ", value).split())


def normalize_ean(value):
    if value is None:
        return ""
    value = str(value).strip()
    if value.endswith(".0"):
        value = value[:-2]
    return re.sub(r"\D", "", value)


def number_from_text(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    text = text.replace("RSD", "").replace("rsd", "")
    text = text.replace("DIN", "").replace("din", "")
    text = text.replace(" ", "")
    if re.fullmatch(r"\d{1,3}(\.\d{3})+", text):
        text = text.replace(".", "")
    elif re.fullmatch(r"\d{1,3}(,\d{3})+", text):
        text = text.replace(",", "")
    else:
        text = text.replace(",", ".")
    try:
        return float(text)
    except Exception:
        return None


def format_price(value):
    if value is None:
        return ""
    value = round(float(value), 2)
    return int(value) if value.is_integer() else value


def get_xml_value(product, tag):
    element = product.find(tag)
    return clean_text(element.text) if element is not None else ""


def download_xml():
    print("PREUZIMANJE TTE ETAIL XML...")
    response = requests.get(
        ETAIL_URL,
        timeout=120,
        headers={"User-Agent": "Mozilla/5.0"},
    )
    print("HTTP status:", response.status_code)
    response.raise_for_status()
    if not response.content:
        raise Exception("TTE XML je prazan.")
    ET.fromstring(response.content)
    print("TTE XML JE ISPRAVAN.")
    return response.content


def parse_tte_xml(xml_data):
    root = ET.fromstring(xml_data)
    products = {}

    for product in root.findall(".//product"):
        ean = normalize_ean(get_xml_value(product, "ean"))
        if not ean:
            continue

        brand = get_xml_value(product, "brand")
        if normalize_text(brand) in EXCLUDED_BRANDS:
            continue

        article_number = get_xml_value(product, "article_number")
        name = get_xml_value(product, "name")

        net_text = get_xml_value(product, "neto_price")
        if not net_text:
            net_text = get_xml_value(product, "net_price")
        if not net_text:
            net_text = get_xml_value(product, "price")

        net_price = number_from_text(net_text)
        if net_price is None:
            continue

        if ean not in products:
            products[ean] = {
                "ean": ean,
                "brand": brand,
                "article_number": article_number,
                "name": name,
                "net_price": net_price,
                "price_with_vat": net_price * PDV,
            }

    print("TTE proizvoda sa EAN-om nakon filtera:", len(products))
    return products


def find_btm_header(sheet):
    best_row = None
    best_score = 0
    for row_index in range(min(sheet.nrows, 50)):
        row_text = " | ".join(
            clean_text(sheet.cell_value(row_index, col)).lower()
            for col in range(sheet.ncols)
        )
        score = 0
        if "ean" in row_text or "barkod" in row_text:
            score += 3
        if "naziv" in row_text:
            score += 3
        if "cena" in row_text:
            score += 2
        if "sifra" in row_text or "šifra" in row_text:
            score += 2
        if score > best_score:
            best_score = score
            best_row = row_index
    if best_row is None:
        raise Exception("Ne mogu da pronađem zaglavlje BTM tabele.")
    return best_row


def find_column(headers, possible_names):
    normalized = [normalize_text(h) for h in headers]
    for i, header in enumerate(normalized):
        for name in possible_names:
            if header == normalize_text(name):
                return i
    for i, header in enumerate(normalized):
        for name in possible_names:
            if normalize_text(name) in header:
                return i
    return None


def read_btm_xls():
    if not os.path.exists(BTM_FILE):
        raise Exception(f"BTM fajl ne postoji: {BTM_FILE}")

    workbook = xlrd.open_workbook(BTM_FILE, formatting_info=False)
    sheet = workbook.sheet_by_index(0)
    header_row = find_btm_header(sheet)
    headers = [clean_text(sheet.cell_value(header_row, c)) for c in range(sheet.ncols)]

    ean_col = find_column(headers, ["EAN", "Barkod", "Bar code", "Barcode", "EAN kod"])
    name_col = find_column(headers, ["Naziv", "Naziv artikla", "Artikal naziv", "Product name", "Name"])
    price_col = find_column(headers, ["Cena", "Cena RSD", "Cena sa PDV", "Price"])
    code_col = find_column(headers, ["Šifra", "Sifra", "Artikal", "Artikal šifra", "Product code", "Code"])

    if ean_col is None:
        raise Exception("Ne mogu da pronađem BTM EAN/Barkod kolonu.")
    if name_col is None:
        raise Exception("Ne mogu da pronađem BTM kolonu sa nazivom artikla.")
    if price_col is None:
        raise Exception("Ne mogu da pronađem BTM kolonu sa cenom.")

    products = []
    for row in range(header_row + 1, sheet.nrows):
        ean = normalize_ean(sheet.cell_value(row, ean_col))
        name = clean_text(sheet.cell_value(row, name_col))
        price = number_from_text(sheet.cell_value(row, price_col))
        if not name or price is None:
            continue
        products.append({
            "ean": ean,
            "name": name,
            "price_with_vat": price,
            "code": clean_text(sheet.cell_value(row, code_col)) if code_col is not None else "",
        })

    print("BTM proizvoda:", len(products))
    return products


def main():
    print("TTE / BTM POREĐENJE")
    print("SAMO EAN | SAMO RAZLIKE | BTM SA PDV | TTE NETO + 20% PDV")

    if not TTE_ETAIL_API_KEY:
        raise Exception("Nedostaje GitHub Secret: TTE_ETAIL_API_KEY")

    btm_products = read_btm_xls()
    tte_products = parse_tte_xml(download_xml())

    results = []

    for btm in btm_products:
        ean = btm["ean"]
        if not ean:
            continue

        tte = tte_products.get(ean)
        if not tte:
            continue

        btm_price = btm["price_with_vat"]
        tte_price = tte["price_with_vat"]
        difference = round(tte_price - btm_price, 2)

        # U tabelu ulaze ISKLJUČIVO artikli sa razlikom u ceni.
        if abs(difference) < 0.01:
            continue

        difference_percent = round((difference / btm_price) * 100, 2) if btm_price else None

        results.append([
            ean,
            tte["brand"],
            btm["code"],
            btm["name"],
            format_price(btm_price),
            tte["article_number"],
            tte["name"],
            format_price(tte["net_price"]),
            format_price(tte_price),
            format_price(difference),
            format_price(difference_percent),
        ])

    # Najveća razlika prva.
    results.sort(key=lambda r: abs(float(r[9])), reverse=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "BTM vs TTE"

    headers = [
        "EAN",
        "Brend",
        "BTM šifra",
        "BTM naziv",
        "BTM cena sa PDV",
        "TTE šifra",
        "TTE naziv",
        "TTE neto",
        "TTE cena sa PDV",
        "Razlika",
        "Razlika %",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    for row in results:
        sheet.append(row)

    widths = [16, 18, 18, 42, 18, 18, 42, 16, 18, 14, 12]
    for i, width in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = width

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    for row in sheet.iter_rows(min_row=2):
        for col in [5, 8, 9, 10, 11]:
            row[col - 1].number_format = '#,##0.00'

    workbook.save(OUTPUT_FILE)
    print("ARTIKALA SA RAZLIKOM:", len(results))
    print("NAPRAVLJEN:", OUTPUT_FILE)


if __name__ == "__main__":
    main()
