import os
import re
import html
import zipfile
import io
import xlrd
import requests
import xml.etree.ElementTree as ET
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

BTM_FILE = "BTM_export.xls"
OUTPUT_FILE = "BTM_TTE_poredjenje.xlsx"

# Ovi brendovi se NE prikazuju u poređenju.
# JBL, Samsung i Apple OSTAJU u poređenju.
EXCLUDED_BRANDS = {
    "bavin", "bavitel", "havit",
    "powerology", "green lion", "porodo"
}

API_KEY = os.environ.get("TTE_ETAIL_API_KEY")
URL = "https://tte.rs/api/sr/products?output_type=xml&api_key=" + str(API_KEY)


def clean(v):
    if v is None:
        return ""
    v = html.unescape(str(v))
    v = re.sub(r"<[^>]+>", " ", v)
    v = re.sub(r"\s+", " ", v.replace("\xa0", " "))
    return v.strip()


def norm(v):
    v = clean(v).lower()
    for a, b in {"č":"c", "ć":"c", "š":"s", "ž":"z", "đ":"d"}.items():
        v = v.replace(a, b)
    return " ".join(re.sub(r"[^a-z0-9]+", " ", v).split())


def ean(v):
    s = clean(v)
    if s.endswith(".0"):
        s = s[:-2]
    return re.sub(r"\D", "", s)


def num(v):
    if v is None or str(v).strip() == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = clean(v).replace("RSD", "").replace("rsd", "").replace("DIN", "").replace("din", "").replace(" ", "")
    if re.fullmatch(r"\d{1,3}(\.\d{3})+", s):
        s = s.replace(".", "")
    elif re.fullmatch(r"\d{1,3}(,\d{3})+", s):
        s = s.replace(",", "")
    else:
        s = s.replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def xml_value(p, tag):
    x = p.find(tag)
    return clean(x.text) if x is not None else ""


def column(headers, names):
    h = [norm(x) for x in headers]
    n = [norm(x) for x in names]
    for i, x in enumerate(h):
        if x in n:
            return i
    for i, x in enumerate(h):
        if any(y in x for y in n):
            return i
    return None


def _rows_from_values(values):
    header = None
    score_best = 0
    for r, row in enumerate(values[:50]):
        text = " | ".join(clean(x).lower() for x in row)
        score = (3 if "ean" in text or "barkod" in text else 0) + (3 if "naziv" in text else 0) + (2 if "cena" in text else 0) + (2 if "sifra" in text or "šifra" in text else 0)
        if score > score_best:
            score_best, header = score, r

    if header is None:
        raise Exception("Ne mogu da pronađem BTM zaglavlje.")

    headers = [clean(x) for x in values[header]]
    ec = column(headers, ["EAN", "Barkod", "Barcode", "EAN kod"])
    nc = column(headers, ["Naziv", "Naziv artikla", "Product name", "Name"])
    pc = column(headers, ["Cena", "Cena RSD", "Cena sa PDV", "Price"])
    cc = column(headers, ["Šifra", "Sifra", "Artikal", "Product code", "Code"])

    if ec is None or nc is None or pc is None:
        raise Exception("Ne mogu da pronađem EAN, naziv ili cenu u BTM fajlu.")

    out = []
    for row in values[header + 1:]:
        row = list(row) + [""] * max(0, len(headers) - len(row))
        code = clean(row[cc]) if cc is not None else ""
        name = clean(row[nc])
        price = num(row[pc])
        code_ean = ean(row[ec])
        if name and price is not None and code_ean:
            out.append({"ean": code_ean, "name": name, "code": code, "price": price})
    return out


def read_btm():
    # BTM portal trenutno vraća XLSX sadržaj sa .xls ekstenzijom.
    # Zato prvo proveravamo ZIP/XLSX format; ako nije XLSX, koristimo pravi XLS parser.
    data = open(BTM_FILE, "rb").read()

    if zipfile.is_zipfile(io.BytesIO(data)):
        wb = load_workbook(io.BytesIO(data), data_only=True, read_only=True)
        sh = wb.active
        values = [list(row) for row in sh.iter_rows(values_only=True)]
        out = _rows_from_values(values)
        wb.close()
    else:
        wb = xlrd.open_workbook(BTM_FILE, formatting_info=False)
        sh = wb.sheet_by_index(0)
        values = [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
        out = _rows_from_values(values)

    print("BTM artikala sa EAN-om:", len(out))
    return out


def read_tte():
    r = requests.get(URL, timeout=120, headers={"User-Agent": "Mozilla/5.0"})
    r.raise_for_status()
    root = ET.fromstring(r.content)
    out = {}

    for p in root.findall(".//product"):
        code_ean = ean(xml_value(p, "ean"))
        if not code_ean:
            continue
        brand = xml_value(p, "brand")
        if norm(brand) in EXCLUDED_BRANDS:
            continue

        price_text = xml_value(p, "neto_price") or xml_value(p, "net_price") or xml_value(p, "price")
        price = num(price_text)
        if price is None:
            continue

        if code_ean not in out:
            out[code_ean] = {
                "brand": brand,
                "code": xml_value(p, "article_number"),
                "name": xml_value(p, "name"),
                "price": price,
            }

    print("TTE artikala sa EAN-om posle filtera:", len(out))
    return out


def main():
    if not API_KEY:
        raise Exception("Nedostaje GitHub Secret: TTE_ETAIL_API_KEY")

    btm = read_btm()
    tte = read_tte()
    results = []

    for b in btm:
        t = tte.get(b["ean"])
        if not t:
            continue

        # OBE CENE SU NETO, BEZ PDV-a. NEMA MNOŽENJA/DELJENJA SA 1.20.
        diff = round(t["price"] - b["price"], 2)
        if abs(diff) < 0.01:
            continue

        pct = round(diff / b["price"] * 100, 2) if b["price"] else None
        results.append([
            b["ean"], t["brand"], b["code"], b["name"], b["price"],
            t["code"], t["name"], t["price"], diff, pct
        ])

    results.sort(key=lambda x: abs(float(x[8])), reverse=True)

    wb = Workbook()
    sh = wb.active
    sh.title = "BTM vs TTE"
    headers = [
        "EAN", "Brend", "BTM šifra", "BTM naziv", "BTM cena NETO",
        "TTE šifra", "TTE naziv", "TTE cena NETO", "Razlika", "Razlika %"
    ]
    sh.append(headers)
    for c in sh[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="center")
    for row in results:
        sh.append(row)

    widths = [16, 18, 18, 42, 18, 18, 42, 18, 14, 12]
    for i, w in enumerate(widths, 1):
        sh.column_dimensions[get_column_letter(i)].width = w
    sh.freeze_panes = "A2"
    sh.auto_filter.ref = sh.dimensions
    for row in sh.iter_rows(min_row=2):
        for i in [5, 8, 9, 10]:
            row[i-1].number_format = '#,##0.00'

    wb.save(OUTPUT_FILE)
    print("ARTIKALA SA RAZLIKOM:", len(results))
    print("NAPRAVLJEN:", OUTPUT_FILE)


if __name__ == "__main__":
    main()
